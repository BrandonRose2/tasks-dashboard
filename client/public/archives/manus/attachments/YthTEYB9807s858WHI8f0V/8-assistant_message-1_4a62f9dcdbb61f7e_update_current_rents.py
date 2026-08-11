"""
update_current_rents.py
-----------------------
Reads rent roll files from the OneDrive folder (mounted at /mnt/desktop/Loss to Lease Analysis/),
extracts average market rent by bedroom type for each property, and populates
CurrentRents_DataEntry.xlsx with the values.

The script is designed to be re-run as new rent rolls are added to the folder.
It only fills in cells that are currently empty (None) and never overwrites existing values.
"""

import os
import glob
import shutil
import xlrd
import openpyxl
from collections import defaultdict
import re
import datetime

# ── Paths ──────────────────────────────────────────────────────────────────────
ONEDRIVE_DIR   = "/mnt/desktop/Loss to Lease Analysis"
SANDBOX_CACHE  = "/home/ubuntu/rent_rolls"
SPREADSHEET_IN = "/home/ubuntu/upload/CurrentRents_DataEntry.xlsx"
SPREADSHEET_OUT = "/home/ubuntu/CurrentRents_DataEntry_updated.xlsx"

os.makedirs(SANDBOX_CACHE, exist_ok=True)

# ── Bedroom-type mapping helpers ───────────────────────────────────────────────
# Maps floorplan codes / keywords → canonical bedroom label used in the spreadsheet
# Patterns are checked in order; first match wins.
FLOORPLAN_BD_MAP = [
    # Exact / prefix patterns for Yardi "Rent Roll Detail" format
    # (Arbor Crest, Boca Ciega, Coral Village, Fairfax, Jefferson, Marrero)
    (re.compile(r'1BR|1X1|1 [A-Z]|1MKT|1MR|^1[A-Z]$', re.I),  '1BD'),
    (re.compile(r'2BR|2X1|2X2|2 [A-Z]|2MKT|2MR|^2[A-Z]$', re.I), '2BD'),
    (re.compile(r'3BR|3X2|3 [A-Z]|3MKT|3MR|^3[A-Z]$|AMS 3|CNRY 3|HKO 3|RBN 3|SS 3', re.I), '3BD'),
    (re.compile(r'4BR|4X3|4 [A-Z]|4MKT|4MR|^4[A-Z]$|AMS 4|CNRY 4|HKO 4|RBN 4|SS 4', re.I), '4BD'),
    (re.compile(r'5BR|5X', re.I), '5BD'),
    # Gates of Manhattan unit-type codes  (gm1x1s → 1BD, gm2x1s/gm2x1swd/gm2x1wd → 2BD, gm3x1.5w → 3BD)
    (re.compile(r'gm1x1', re.I), '1BD'),
    (re.compile(r'gm2x1', re.I), '2BD'),
    (re.compile(r'gm3x', re.I),  '3BD'),
    # Ruby Diamond unit-type codes
    # dia101/rub101 (~1503 sf) → 3BD, dia102/ru102 (~1584 sf) → 4BD
    (re.compile(r'dia101|rub101', re.I), '3BD'),
    (re.compile(r'dia102|ru102', re.I), '4BD'),
    # Star unit-type codes
    # sr100 (~1001 sf) → 2BD, sr101 (~1192 sf) → 3BD, sr102 (~1452 sf) → 4BD
    (re.compile(r'sr100', re.I), '2BD'),
    (re.compile(r'sr101', re.I), '3BD'),
    (re.compile(r'sr102', re.I), '4BD'),
    # Jefferson floorplan codes: 2A → 2BD, 3A → 3BD
    (re.compile(r'^2A$', re.I), '2BD'),
    (re.compile(r'^3A$', re.I), '3BD'),
    # Boca Ciega: 2B/2BFLAT → 2BD, 3BFLAT → 3BD
    (re.compile(r'2B|2BFLAT', re.I), '2BD'),
    (re.compile(r'3BFLAT', re.I), '3BD'),
]

def floorplan_to_bd(fp_code: str) -> str | None:
    """Return '1BD', '2BD', etc. or None if unrecognised."""
    fp = str(fp_code).strip()
    for pattern, label in FLOORPLAN_BD_MAP:
        if pattern.search(fp):
            return label
    return None


# ── Property → folder mapping ──────────────────────────────────────────────────
# Maps the property name as it appears in the spreadsheet to the OneDrive subfolder name.
PROPERTY_FOLDER_MAP = {
    '21 - Boca Ciega':       'Boca Ciega',
    '13 - Jefferson':        'Jefferson',
    '18 - Opa Locka':        'Opa Locka',
    '14 - Macedonia':        'Macedonia',
    '48 - Coral Village':    'Coral Village',
    '15 - Holiday':          'Holiday',
    '19 - Lexington':        'Lexington',
    '20 - La Promesa':       'La Promesa',
    '22 - Thibodaux':        'Thibodaux',
    '23 - Breckenridge':     'Breckenridge',
    '29 - Grace Townhomes':  'Grace Townhomes',
    '30 - River Pointe':     'River Pointe',
    '32 - Walnut Hill':      'Walnut Hill',
    '35 - Cumberland':       'Cumberland',
    '38 - Grove Park':       'Grove Park',
    '19 - Gates':            'Gates of Manhattan',
    '40 - Marrero':          'Marrero',
    '37 - Arbor Crest':      'Arbor Crest',
    '33 - Windsor-Yorkshire': 'Windsor-Yorkshire',
    '50 - North Pointe':     'North Pointe',
    '53 - Bayou Pointe':     'Bayou Pointe',
    '41 - St. Charles':      'St. Charles',
    '55 - Pelican Bay':      'Pelican Bay',
    '56 - Howell Place':     'Howell Place',
    '57 - Pirates Bend':     'Pirates Bend',
    '6 - Anaheim Gardens':   'Anaheim Gardens',
    '11 - Fairfax':          'Fairfax',
    '31 - Urban':            'Urban',
    '2 - Midtown':           'Midtown',
    '34 - Pacific Pointe':   'Pacific',
    '28 - Granite Ridge':    'Granite Ridge',
    '25 - Columbia':         'Columbia',
    '26 - Forest View':      'Forest View',
    '24 - Oak Hills':        'Oak Hills',
    '58 - River Garden':     'River Garden',
    '36 - Silver Springs':   'Silver Springs',
    '44 - Thomasville':      'Thomasville',
    '63 - Crossroads':       'Crossroads',
    '42 - Ruby':             'Ruby Diamond',
    '39 - Star':             'Star',
}


# ── File discovery ─────────────────────────────────────────────────────────────
def find_rent_rolls_for_property(folder_name: str) -> list[str]:
    """Return list of rent roll file paths (copied to sandbox cache) for a property folder."""
    prop_dir = os.path.join(ONEDRIVE_DIR, folder_name)
    if not os.path.isdir(prop_dir):
        return []
    results = []
    for fname in os.listdir(prop_dir):
        if fname.startswith('.'):
            continue
        if fname.lower().endswith(('.xls', '.xlsx')):
            src = os.path.join(prop_dir, fname)
            dst = os.path.join(SANDBOX_CACHE, fname)
            if not os.path.exists(dst):
                shutil.copy2(src, dst)
            results.append(dst)
    return results


# ── Extraction: Yardi "Rent Roll Detail" format (Sheet2 summary) ───────────────
def extract_yardi_summary(filepath: str) -> dict[str, float]:
    """
    Parse the Sheet2 summary tab of a Yardi Rent Roll Detail export.
    Returns {bd_label: avg_market_rent} e.g. {'2BD': 1020.0, '3BD': 1150.0}
    """
    result = {}
    try:
        if filepath.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filepath)
            sheet_names = wb.sheetnames
        else:
            wb = xlrd.open_workbook(filepath, logfile=open(os.devnull, 'w'))
            sheet_names = wb.sheet_names()

        # Look for Sheet2 (summary)
        summary_sheet = None
        for sn in sheet_names:
            if 'Sheet2' in sn or 'Summary' in sn.lower():
                summary_sheet = sn
                break
        if not summary_sheet:
            return result

        if filepath.endswith('.xlsx'):
            ws = wb[summary_sheet]
            rows = [tuple(ws.cell(r, c).value for c in range(1, ws.max_column + 1))
                    for r in range(1, ws.max_row + 1)]
        else:
            ws = wb.sheet_by_name(summary_sheet)
            rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

        # Find header row (contains 'Floorplan')
        header_row_idx = None
        for i, row in enumerate(rows):
            if any('Floorplan' in str(v) for v in row if v):
                header_row_idx = i
                break
        if header_row_idx is None:
            return result

        # Data rows follow header; skip "Totals" row
        bd_rents = defaultdict(list)
        for row in rows[header_row_idx + 1:]:
            fp = row[0] if row else None
            if not fp or 'Total' in str(fp):
                continue
            # Average Market + Addl. is column index 3 (0-based)
            avg_market = row[3] if len(row) > 3 else None
            if not avg_market or not isinstance(avg_market, (int, float)) or avg_market == 0:
                # Fall back to Average Leased (col 5)
                avg_market = row[5] if len(row) > 5 else None
            if not avg_market or not isinstance(avg_market, (int, float)):
                continue
            bd = floorplan_to_bd(str(fp))
            if bd:
                bd_rents[bd].append(float(avg_market))

        for bd, rents in bd_rents.items():
            result[bd] = round(sum(rents) / len(rents))

    except Exception as e:
        print(f"  [WARN] extract_yardi_summary({os.path.basename(filepath)}): {e}")
    return result


# ── Extraction: "Rent Roll with Lease Charges" format (Gates, Ruby Diamond, Star) ──
def extract_lease_charges_format(filepath: str) -> dict[str, float]:
    """
    Parse the single-sheet 'Rent Roll with Lease Charges' format used by
    Gates of Manhattan, Ruby Diamond, and Star.
    Returns {bd_label: avg_market_rent}
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        unit_type_rents = defaultdict(list)
        for row in ws.iter_rows(values_only=True):
            # Main data row: col[0]=unit_id (str), col[1]=unit_type, col[5]=market_rent (num)
            if (row[0] and row[1] and row[5]
                    and isinstance(row[0], str)
                    and isinstance(row[1], str)
                    and isinstance(row[5], (int, float))
                    and row[5] > 0):
                unit_type_rents[row[1]].append(float(row[5]))

        bd_rents = defaultdict(list)
        for ut, rents in unit_type_rents.items():
            bd = floorplan_to_bd(ut)
            if bd:
                bd_rents[bd].extend(rents)

        for bd, rents in bd_rents.items():
            result[bd] = round(sum(rents) / len(rents))

    except Exception as e:
        print(f"  [WARN] extract_lease_charges_format({os.path.basename(filepath)}): {e}")
    return result


# ── Detect format and extract ──────────────────────────────────────────────────
def extract_rents_from_file(filepath: str) -> dict[str, float]:
    """Auto-detect format and return {bd_label: avg_market_rent}."""
    fname = os.path.basename(filepath).lower()
    # Lease-charges format is always .xlsx and has no Sheet2
    if filepath.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filepath)
        if len(wb.sheetnames) == 1 and 'Report' in wb.sheetnames[0]:
            return extract_lease_charges_format(filepath)
        else:
            return extract_yardi_summary(filepath)
    else:
        return extract_yardi_summary(filepath)


# ── Main logic ─────────────────────────────────────────────────────────────────
def build_property_rents() -> dict[str, dict[str, float]]:
    """
    Scan all property folders and return:
      { property_name_in_spreadsheet: { '1BD': 1200, '2BD': 1500, ... } }
    Tries all files in the folder and merges results, preferring files with more bedroom types.
    """
    property_rents = {}
    for prop_name, folder_name in PROPERTY_FOLDER_MAP.items():
        files = find_rent_rolls_for_property(folder_name)
        if not files:
            continue
        # Try all files and pick the one with the most bedroom types
        best_rents = {}
        for f in files:
            print(f"  Processing [{prop_name}] → {os.path.basename(f)}")
            rents = extract_rents_from_file(f)
            if len(rents) > len(best_rents):
                best_rents = rents
                print(f"    Extracted: {rents}")
            else:
                print(f"    Extracted (not best): {rents}")
        if best_rents:
            property_rents[prop_name] = best_rents
        else:
            print(f"    No data extracted for [{prop_name}].")
    return property_rents


def update_spreadsheet(property_rents: dict[str, dict[str, float]]):
    """
    Load the spreadsheet, fill in empty Current Rent cells, and save.
    Only fills cells where the value is currently None/empty.
    """
    wb = openpyxl.load_workbook(SPREADSHEET_IN)
    ws = wb.active

    filled = 0
    skipped_existing = 0
    not_found = 0

    for row_idx in range(3, ws.max_row + 1):  # data starts at row 3
        prop_cell  = ws.cell(row_idx, 1).value
        bd_cell    = ws.cell(row_idx, 2).value
        rent_cell  = ws.cell(row_idx, 3).value

        if not prop_cell or not bd_cell:
            continue
        if rent_cell is not None:
            skipped_existing += 1
            continue

        prop_name = str(prop_cell).strip()
        bd_label  = str(bd_cell).strip()

        if prop_name not in property_rents:
            not_found += 1
            continue

        bd_rents = property_rents[prop_name]
        if bd_label not in bd_rents:
            not_found += 1
            continue

        ws.cell(row_idx, 3).value = bd_rents[bd_label]
        filled += 1

    print(f"\nSpreadsheet update summary:")
    print(f"  Cells filled:          {filled}")
    print(f"  Cells already had data:{skipped_existing}")
    print(f"  Cells with no match:   {not_found}")

    wb.save(SPREADSHEET_OUT)
    print(f"\nSaved updated spreadsheet to: {SPREADSHEET_OUT}")


if __name__ == '__main__':
    print(f"=== update_current_rents.py  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
    print("Scanning property folders for rent rolls...\n")
    property_rents = build_property_rents()
    print(f"\nFound rent data for {len(property_rents)} properties.")
    update_spreadsheet(property_rents)
